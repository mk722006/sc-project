def local_report(month, year, selected_country, table_name, conn, curr):
    
    # Imports
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image #install Pillow 

    import pandas as pd
    from datetime import datetime, timedelta

    import shutil
    import os
    from dateutil.relativedelta import relativedelta

    # MTD, YTD - get_date_range | Dynamic Date Range - Function 

    def get_date_range(month, year):
        # Calculate the starting date range
        if month == 1:
            starting_date = datetime(year=year - 1, month=12, day=31)
        else:
            starting_date = datetime(year=year, month=month, day=1) - timedelta(days=1)

        # Calculate the ending date range
        if month == 12:
            ending_date = datetime(year=year + 1, month=1, day=1)
        else:
            ending_date = datetime(year=year, month=month + 1, day=1)

        # Format the dates as strings
        starting_date_str = starting_date.strftime('%Y-%m-%d')
        ending_date_str = ending_date.strftime('%Y-%m-%d')
        
        starting_date_ytd = datetime(year=year - 1, month=12, day=31)
        starting_date_ytd_str = starting_date_ytd.strftime('%Y-%m-%d')

        return starting_date_str, ending_date_str, starting_date_ytd_str, ending_date_str

    def convert_month_year(month, year):
        # Get the month abbreviation
        month_abbr = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        month_str = month_abbr[month]

        # Get the last two digits of the year
        year_str = str(year) #[-2:]

        # Combine the month abbreviation and year
        output = f"{month_str}-{year_str}"
        return output
    ## Important Variables
    start_date_mtd, end_date_mtd, start_date_ytd, end_date_ytd= get_date_range(month, year)

    ## All Market Info - market_df
    markets_loop = ['Global','Brazil','Chile','Colombia', 'Peru']
    currency_loop = ['USD','BRL','CLP','COP','PEN']
    curr_symbol_loop = ['$','R$','$','$','S/']
    market_df = pd.DataFrame({'country': markets_loop, 'currency': currency_loop, 'currency_symbol': curr_symbol_loop})


    ## Sheet Name, Currency Symbol, Country
    month_year_abbr = convert_month_year(month, year)
    sheet_name = f'{selected_country} Exe Sum' # Ex: Brazil Exe Sum
    curr_symbol = market_df[market_df['country']==selected_country]['currency'].values[0]
    print(f'Country: {selected_country} | Sheet Name: {sheet_name} | Symbol: {curr_symbol} | Month: {month_year_abbr}')


    print(f'Starting Date MTD: {start_date_mtd} | Ending Date MTD: {end_date_mtd}')
    print(f'Starting Date YTD: {start_date_ytd} | Ending Date YTD: {end_date_ytd}')
    ## Making a Copy of the Template File | Different for Global and Other countries
    file_name = f"executive_report-{selected_country}-{convert_month_year(month, year).split('-')[0]}-{year}.xlsx"

    if selected_country == 'Global':
        source_file_path = 'Template/Template - Global Exe Sum.xlsx'    
    else: # Brazil, Chile, Colombia, Peru
        source_file_path = 'Template/Template - Others.xlsx'
        
    # Create a copy file path
    copy_directory = 'Copies'
    os.makedirs(copy_directory, exist_ok=True)
    copy_file_path = f'{copy_directory}/{file_name}'

    try: ## Removing the file if already available
        os.remove(copy_file_path)
    except:
        pass


    # Open the copied file
    copy_wb = load_workbook(source_file_path)

    # Deleting Unwanted Sheets from - Others Template
    if selected_country!= 'Global':
        for name in copy_wb.sheetnames:
            if name != sheet_name:
                copy_wb.remove(copy_wb[name])
    # Required Functions
    def calculate_previous_dates(month, year):
        # Create a datetime object for the first day of the given month and year
        date = datetime(year, month, 1)
        
        # Future date
        future_date = date - relativedelta(months=-1)
        current_date = date - relativedelta(months=0)
        one_month_prior = date - relativedelta(months=1)
        two_months_prior = date - relativedelta(months=2)
        three_months_prior = date - relativedelta(months=3)
        
        ### Month Names ####
        three_months_name = convert_month_year(three_months_prior.month, three_months_prior.year)
        two_months_name = convert_month_year(two_months_prior.month, two_months_prior.year)
        one_months_name = convert_month_year(one_month_prior.month, one_month_prior.year)
        current_month_name = convert_month_year(current_date.month, current_date.year)
        
        
        ## Start and End Date ##
        three_months_prior_start_date = three_months_prior - timedelta(days=1)
        three_months_prior_start_date = three_months_prior_start_date.strftime('%Y-%m-%d')
        three_months_prior_end_date = two_months_prior.strftime('%Y-%m-%d')

        two_months_prior_start_date = two_months_prior - timedelta(days=1)
        two_months_prior_start_date = two_months_prior_start_date.strftime('%Y-%m-%d')
        two_months_prior_end_date = one_month_prior.strftime('%Y-%m-%d')


        one_month_prior_start_date = one_month_prior - timedelta(days=1)
        one_month_prior_start_date = one_month_prior_start_date.strftime('%Y-%m-%d')
        one_month_prior_end_date = current_date.strftime('%Y-%m-%d')

        current_date_start_date = current_date - timedelta(days=1)
        current_date_start_date = current_date_start_date.strftime('%Y-%m-%d')
        current_date_end_date = future_date.strftime('%Y-%m-%d')

        return [[three_months_prior_start_date, three_months_prior_end_date, three_months_name], [two_months_prior_start_date, two_months_prior_end_date, two_months_name], [one_month_prior_start_date, one_month_prior_end_date, one_months_name], [current_date_start_date, current_date_end_date, current_month_name]]

    def getting_all_others_info(dictionary_info):
        df = pd.DataFrame(dictionary_info)
        percent_100 = (df[df.columns[1]].sum() / (df[df.columns[2]].sum())) * 100
        all_others = round(percent_100 - df[df.columns[1]].sum(), 2)
        all_others_percentage = round((100 - df[df.columns[2]].sum())/100, 3)
        return all_others, all_others_percentage

    def spend_percentage_value_function(dictionary_data, column_for_name, column_spend, column_percentage, copy_wb, cell_number, sheet_name, run):
        
        if column_for_name!=False: ## Set to False when you do not have column_for_name
            cell_name_A = f'{column_for_name}{cell_number}'
            cell_value_spend = list(dictionary_data[run].values())[0] # Name 
            copy_wb[sheet_name][cell_name_A].value = cell_value_spend
            

        if column_spend!= False:
        
            cell_name_B = f'{column_spend}{cell_number}'
            cell_value_spend = list(dictionary_data[run].values())[1] # Spend Value 
            copy_wb[sheet_name][cell_name_B].value = round(cell_value_spend, 2)
        
        if column_percentage!= False:
            cell_name_C = f'{column_percentage}{cell_number}'
            cell_value_percentage = round(list(dictionary_data[run].values())[2]/100, 3) # Percentage Value | Dividing by 100
            copy_wb[sheet_name][cell_name_C].value = cell_value_percentage
            

    three_months_prior, two_months_prior, one_month_prior, current_date = calculate_previous_dates(month, year)
    print(three_months_prior, two_months_prior, one_month_prior, current_date )
    # =_=_==_=_==_=_==_=_==_=_= {0} Some Headers - Other Countries =_=_==_=_==_=_==_=_==_=_=
    ## A2 | Ex: Global Supply Chain Analytics - Apr 2023
    copy_wb[sheet_name]['A2'].value = f'Global Supply Chain Analytics - {month_year_abbr.split("-")[0]} 2023'


    ## B6, D6 | Ex: CLP Spend MTD
    copy_wb[sheet_name]['B6'].value = f'{curr_symbol} Spend MTD'
    copy_wb[sheet_name]['D6'].value = f'{curr_symbol} Spend YTD'

    ## MTD Currency Headers: E47, E98, E152 | Ex: MTD BRL Spend
    mtd_currency_spend_headers = ['E47', 'E98', 'E152']
    for head in mtd_currency_spend_headers:
        copy_wb[sheet_name][head].value = f'MTD {curr_symbol} Spend'
        
    ## YTD Currency Headers: F47, F98, F152 | Ex: YTD BRL Spend
    ytd_currency_spend_headers = ['F47', 'F98', 'F152']
    for head in ytd_currency_spend_headers:
        copy_wb[sheet_name][head].value = f'YTD {curr_symbol} Spend'

    ## Headers: B71, E71, H71, K71, B86, E86, H86, K86, B124, E124, H124, K124, B140, E140, H140, K140, B164, E164, H164, K164, B179, E179, H179, K179 | Ex: BRL
    headers_currency_list = ['B71','E71','H71','K71','B86','E86','H86','K86','B124','E124','H124','K124','B140','E140','H140','K140','B164','E164','H164','K164','B179','E179','H179','K179']
    for head in headers_currency_list:
        copy_wb[sheet_name][head].value = curr_symbol
    # =_=_==_=_==_=_==_=_==_=_= {1} Executive Summary  [ROW-4] =_=_==_=_==_=_==_=_==_=_=

    ### **{1.1} USD Spend MTD [ROW-6]**

    ### **{1.2} USD Spend YTD [ROW-6]**

    ### **{1.3} Unique Suppliers [ROW-6]**

    ### **{1.4} Unique Manufacturers [ROW-6]**

    ### **{1.5} Unique SKUs [ROW-6]**
    def convert_to_millions_or_billions(value, selected_country, market_df):
        currency_to_use = market_df[market_df['country']==selected_country]['currency_symbol'].values[0]
        suffixes = ['', 'K', 'M', 'B', 'T']  # Suffixes for Thousand, Million, Billion, Trillion, etc.
        for suffix in suffixes[1:]:
            value /= 1000.0
            if abs(value) < 1000.0:
                return f"{currency_to_use}{value:.2f}{suffix}"
        return f"{currency_to_use}{value:.2f}{suffixes[-1]}"

    def executive_summary_query_values_function(table_name, start_date_mtd, end_date_mtd, start_date_ytd, end_date_ytd, curr):

        ## query_1_1_usd_spend_mtd ## 
        query_1_1_usd_spend_mtd = f"""select sum (original_total_cost) as USD_Spend_MTD from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and geography = '{selected_country}'"""
        curr.execute(query_1_1_usd_spend_mtd)

        value_query_1_1_usd_spend_mtd = curr.fetchall()
        value_query_1_1_usd_spend_mtd = convert_to_millions_or_billions(list(value_query_1_1_usd_spend_mtd[0].values())[0], selected_country, market_df)

        ## query_1_2_usd_spend_ytd ## 
        query_1_2_usd_spend_ytd = f"""select sum (original_total_cost) as USD_Spend_YTD from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and geography = '{selected_country}'"""
        curr.execute(query_1_2_usd_spend_ytd)

        value_query_1_2_usd_spend_ytd = curr.fetchall()
        value_query_1_2_usd_spend_ytd = convert_to_millions_or_billions(list(value_query_1_2_usd_spend_ytd[0].values())[0], selected_country, market_df)

        ## query_1_3_unique_suppliers ## 
        query_1_3_unique_suppliers = f"""select COUNT(DISTINCT(distributor_normalized)) as unique_suppliers from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and geography = '{selected_country}'"""
        curr.execute(query_1_3_unique_suppliers)

        value_query_1_3_unique_suppliers = curr.fetchall()
        # print(value_query_1_3_unique_suppliers)

        ## query_1_4_unique_suppliers ## 
        query_1_4_unique_manufacturers = f"""select COUNT(DISTINCT(mnf_dashboard_half)) as unique_manufacturers from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and geography = '{selected_country}'"""
        curr.execute(query_1_4_unique_manufacturers)

        value_query_1_4_unique_manufacturers = curr.fetchall()
        # print(value_query_1_4_unique_manufacturers)

        ## query_1_5_unique_skus ##
        query_1_5_unique_skus = f"""select COUNT(DISTINCT(sc_uhg_id)) as unique_SKUs from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and geography = '{selected_country}'"""
        curr.execute(query_1_5_unique_skus)

        value_query_1_5_unique_skus = curr.fetchall()
        # print(value_query_1_5_unique_skus)
        
        return value_query_1_1_usd_spend_mtd, value_query_1_2_usd_spend_ytd, list(value_query_1_3_unique_suppliers[0].values())[0], list(value_query_1_4_unique_manufacturers[0].values())[0], list(value_query_1_5_unique_skus[0].values())[0]

    ######## Running Function ########
    executive_summary_values = executive_summary_query_values_function(table_name, start_date_mtd, end_date_mtd, start_date_ytd, end_date_ytd, curr)


    ## Writing the Values on Excel File ##
    cell_names_executive_summary = ['B7', 'D7', 'F7', 'H7', 'J7']
    for j, value in enumerate(executive_summary_values):
        cell_value = value
        copy_wb[sheet_name][cell_names_executive_summary[j]].value = cell_value
    # =_=_==_=_==_=_==_=_==_=_= {2} Total Purchase Order Spend [ROW-17] =_=_==_=_==_=_==_=_==_=_=
    ### Q: Total P.O. Spend - MTD	| B21-B24 | total_po_spend_mtd
    ## {2.1}Total P.O. Spend - MTD	[ROW-19]
    query_mtd = f'''select geography ,sum (original_total_cost) as spend_MTD ,(spend_MTD / SUM(spend_MTD) OVER ()) * 100 AS percentage_spend_MTD from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and geography = '{selected_country}'  GROUP BY 1'''
    curr.execute(query_mtd)
    total_po_spend_mtd = curr.fetchall()


    ## {2.2} Total P.O. Spend - YTD	[ROW-19]
    query_ytd = f"select geography ,sum (original_total_cost) as spend_YTD ,(spend_YTD / SUM(spend_YTD) OVER ()) * 100 AS percentage_spend_YTD from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and geography = '{selected_country}'  GROUP BY 1"
    curr.execute(query_ytd)
    total_po_spend_ytd = curr.fetchall()
    ### B20, D20 - Same Input [Ex:  Apr-23 ( BRL )] | Row 21 | Column: B,C,D,E 
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Total PO Spend')

    # Writing in Cells - B20, D20  | Same for all countries
    B20_D20_Header_cell_value =  f'{month_year_abbr} ( {curr_symbol} )' # B20 and D20 Cell Values are the same
    header_list = ['B20', 'D20']
    for cell_name in header_list:
        copy_wb[sheet_name][cell_name].value = B20_D20_Header_cell_value
        
    # Row 21  #-# Column: B,D
    cell_number = 21 

    ## B:  total_po_spend_mtd | cell_value_spend | B-21
    B_Cell_Value = spend_percentage_value_function(total_po_spend_mtd, False, 'B', False, copy_wb, cell_number, sheet_name, 0)
    ## D total_po_spend_ytd | cell_value_spend | D-21
    D_Cell_Value = spend_percentage_value_function(total_po_spend_ytd, False, 'D', False, copy_wb, cell_number, sheet_name, 0)
    # =_=_==_=_==_=_==_=_==_=_= {3}Total Category Spend by Market [ROW -28] =_=_==_=_==_=_==_=_==_=_=
    # {3.1} Non-Pharma [ROW-30]
    query_3_1_non_pharma = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_3_1_non_pharma)

    total_category_spend_by_market_non_pharma = curr.fetchall()

    # {3.2} Pharma [ROW-30]
    query_3_2_pharma = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_3_2_pharma)

    total_category_spend_by_market_pharma = curr.fetchall()

    # {3.3} Indirect [ROW-30]
    query_3_3_indirect = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_3_3_indirect)

    total_category_spend_by_market_indirect = curr.fetchall()

    # {3.4} Total PO Spend	[ROW-30]
    query_3_4_total_po_spend = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_3_4_total_po_spend)

    total_category_spend_by_market_total_po_spend = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Total Category Spend by Market')

    # Writing in Cells - B26, D26, F26, H26 | Headers
    table_header_value = f'YTD {year} ( {curr_symbol} )'
    table_header_cells = ['B26', 'D26', 'F26', 'H26']
    for cell in table_header_cells:
        copy_wb[sheet_name][cell].value = table_header_value

        
    ####### Row 32 ########
    cell_number = 27
    run = 0 ## Only One Value in the dictionary

    ## B:  total_category_spend_by_market_non_pharma | cell_value_spend | cell_value_percentage | Row 27
    B_Cell_values = spend_percentage_value_function(total_category_spend_by_market_non_pharma, False, 'B', False, copy_wb, cell_number, sheet_name, run)

    ## D:  total_category_spend_by_market_pharma | cell_value_spend | cell_value_percentage | Row 27
    D_Cell_values = spend_percentage_value_function(total_category_spend_by_market_pharma, False, 'D', False, copy_wb, cell_number, sheet_name, run)

    ## F:  total_category_spend_by_market_indirect | cell_value_spend | cell_value_percentage | Row 27
    F_Cell_values = spend_percentage_value_function(total_category_spend_by_market_indirect, False, 'F', False, copy_wb, cell_number, sheet_name, run)

    ## H:  total_category_spend_by_market_total_po_spend | cell_value_spend | cell_value_percentage | Row 27
    H_Cell_values = spend_percentage_value_function(total_category_spend_by_market_total_po_spend, False, 'H', False, copy_wb, cell_number, sheet_name, run)
    # =_=_==_=_==_=_==_=_==_=_= {4} Top Manufacturers by Category Spend [ROW-48]  =_=_==_=_==_=_==_=_==_=_=
    # {4.1} Non-Pharma [ROW-30]
    query_4_1_non_pharma = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 5"""
    curr.execute(query_4_1_non_pharma)

    top_manufacturers_by_category_spend_non_pharma = curr.fetchall()

    # {4.2} Pharma [ROW-50]
    query_4_2_pharma = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 5"""
    curr.execute(query_4_2_pharma)

    top_manufacturers_by_category_spend_pharma = curr.fetchall()

    # {4.3} Indirect-IT [ROW-50]
    query_4_3_indirect = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 5"""
    curr.execute(query_4_3_indirect)

    top_manufacturers_by_category_spend_indirect = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Top Manufacturers by Category Spend')
    # Writing in Cells - 'B36', 'F36', 'J36' | Headers
    table_header_value = f'YTD {year} ( {curr_symbol} )'
    table_header_cells = ['B36', 'F36', 'J36']

    for cell in table_header_cells:
        copy_wb[sheet_name][cell].value = table_header_value

    for run in range(len(top_manufacturers_by_category_spend_non_pharma)):
        ####### Row 37 to 41 ######## 
        cell_number = 37 + run 
        
        ## A, B & C:  top_manufacturers_by_category_spend_non_pharma | cell_value_spend | cell_value_percentage | Row 37 to 41 | A,B,C
        ABC_Cell_values = spend_percentage_value_function(top_manufacturers_by_category_spend_non_pharma, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)
        
        ## E, F & G:  top_manufacturers_by_category_spend_pharma | cell_value_spend | cell_value_percentage | Row 37 to 41 | E,F,G 
        EFG_Cell_values = spend_percentage_value_function(top_manufacturers_by_category_spend_pharma, 'E', 'F', 'G', copy_wb, cell_number, sheet_name, run)
        
        ## I, J & K:  top_manufacturers_by_category_spend_indirect | cell_value_spend | cell_value_percentage | Row 37 to 41 | I,J,K
        IJK_Cell_values = spend_percentage_value_function(top_manufacturers_by_category_spend_indirect, 'I', 'J', 'K', copy_wb, cell_number, sheet_name, run)

    ## Providing All Others Info | B42, F42, J42 ## 
    all_others, all_others_percentage = getting_all_others_info(top_manufacturers_by_category_spend_non_pharma)
    copy_wb[sheet_name]['B42'].value = all_others
    copy_wb[sheet_name]['C42'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(top_manufacturers_by_category_spend_pharma)
    copy_wb[sheet_name]['F42'].value = all_others
    copy_wb[sheet_name]['G42'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(top_manufacturers_by_category_spend_indirect)
    copy_wb[sheet_name]['J42'].value = all_others
    copy_wb[sheet_name]['K42'].value = all_others_percentage
    # =_=_==_=_== {5} Direct (Non-pharma) Spend [ROW-60]  =_=_==_=_==
    # {5.1} Total P.O. Spend - MTD [ROW-62]
    query_5_1_po_spend = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_5_1_po_spend)
    direct_non_pharma_spend_total_po_spend = curr.fetchall()

    # {5.2} MTD USD Spend [ROW-62] -------- No Need ---------- Taken from Excel Formula
    # {5.3} YTD USD Spend [ROW-62]  ------------- No Need ------------ Taken from Excel Formula


    # {5.4} Total P.O. Spend - YTD Trended [ROW -73]  | {5.4} Total P.O. Spend - YTD Trended [ROW -73] 

    ## 5.4.1 April YTD USD [ROW -74]
    query_5_4_1_ytd = f"""select geography ,sum (original_total_cost) as MTD_April_spend ,(MTD_April_spend / SUM(MTD_April_spend) OVER ()) * 100 AS percentage_MTD_April_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_5_4_1_ytd)
    direct_non_pharma_spend_ytd  = curr.fetchall()

    ## 5.4.2 JAN -23 [ROW -74] | three_months_prior  
    starting_date, ending_date = three_months_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_non_pharma_spend_three_months_prior_value = curr.fetchall()


    ## 5.4.3 FEB -23 [ROW-74] |  two_months_prior  
    starting_date, ending_date = two_months_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_non_pharma_spend_two_months_prior_value = curr.fetchall()


    ## 5.4.4 MAR -23 [ROW-74] |  one_month_prior  
    starting_date, ending_date = one_month_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_non_pharma_spend_one_month_prior_value = curr.fetchall()


    ## 5.4.5 APR - 23 [ROW -74] | current_date 
    starting_date, ending_date = current_date[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_non_pharma_spend_current_date_value = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Direct (Non-pharma) Spend')

    # Writing in Cells - B48, B59 | Headers
    copy_wb[sheet_name]['B48'].value = f'{month_year_abbr} ( {curr_symbol} )'  # Ex: Apr-23 ( USD )
    copy_wb[sheet_name]['B59'].value = f"{month_year_abbr.split('-')[0]} YTD {curr_symbol}"  # Ex: Apr YTD USD 

    # Writing in Cells - D59, E59, F59, G59 | Headers 
    copy_wb[sheet_name]['D59'].value = three_months_prior[-1]  # Jan-23
    copy_wb[sheet_name]['E59'].value = two_months_prior[-1]  # Feb-23
    copy_wb[sheet_name]['F59'].value = one_month_prior[-1]  # Mar-23
    copy_wb[sheet_name]['G59'].value = current_date[-1]  # Apr-23



    ####### Row 49 to 52 [One Row] ######## 
    index_to_add = market_df[market_df['country'] == selected_country].index[0] - 1
    cell_number = 49 + index_to_add
    run = 0

    ## B:  direct_non_pharma_spend_total_po_spend | cell_value_spend | cell_value_percentage | Row 49 to 52 [One Row]
    B_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_total_po_spend, False, 'B', False, copy_wb, cell_number, sheet_name, run)


    cell_number = 60 + index_to_add
    ## B:  direct_non_pharma_spend_ytd | cell_value_spend | cell_value_percentage | Row 60 to 63 [One Row]
    B_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_ytd, False, 'B', False, copy_wb, cell_number, sheet_name, run)

    ## D:  direct_non_pharma_spend_three_months_prior_value | cell_value_spend | Row 60 to 63 [One Row]
    D_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_three_months_prior_value, False, 'D', False, copy_wb, cell_number, sheet_name, run)

    ## E:  direct_non_pharma_spend_two_months_prior_value | cell_value_spend | Row 60 to 63 [One Row]
    E_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_two_months_prior_value, False, 'E', False, copy_wb, cell_number, sheet_name, run)

    ## F:  direct_non_pharma_spend_one_month_prior_value | cell_value_spend | Row 60 to 63 [One Row]
    F_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_one_month_prior_value, False, 'F', False, copy_wb, cell_number, sheet_name, run)

    ## G:  direct_non_pharma_spend_current_date_value | cell_value_spend | Row 60 to 63 [One Row]
    G_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_current_date_value, False, 'G', False, copy_wb, cell_number, sheet_name, run)
    # =_=_==_=_== {6} Top Manufacturers | Top Suppliers | Top Product Categories | Direct (Non-Pharma) Spend=_=_==_=_==
    # {6.1} Top Manufacturers  Apr-23  MTD | Row 70
    query_top_manufacturers = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_manufacturers)
    direct_non_pharma_spend_top_manufacturers_value_mtd = curr.fetchall()

    # {6.2} Top Manufacturers  Apr-23  YTD | Row 70
    query_top_manufacturers_ytd = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_manufacturers_ytd)
    direct_non_pharma_spend_top_manufacturers_value_ytd = curr.fetchall()

    # {6.3} Top Suppliers  Apr-23  MTD | Row 70
    query_top_suppliers = f"""select distributor_normalized as suppliers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_suppliers)
    direct_non_pharma_spend_top_suppliers_value_mtd = curr.fetchall()

    # {6.4} Top Suppliers  Apr-23  YTD | Row 70
    query_top_suppliers_ytd = f"""select distributor_normalized as suppliers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_suppliers_ytd)
    direct_non_pharma_spend_top_suppliers_value_ytd = curr.fetchall()

    # {6.5} Top Product Categories  Apr-23  MTD | Row 85
    query_top_product_categories = f"""select unspsc_class_title as categories ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_product_categories)
    direct_non_pharma_spend_top_product_categories_value_mtd = curr.fetchall()

    # {6.6} Top Product Categories  Apr-23  YTD	| Row 85			
    query_top_product_categories_ytd = f"""select unspsc_class_title as categories ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_product_categories_ytd)
    direct_non_pharma_spend_top_product_categories_value_ytd = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Direct (Non-pharma) Spend | Top Suppliers | Top Manufacturers | Top Product Categories')

    # Writing in Cells - A70, D70, G70, J70, A85, G85 | Headers | Suppliers, Manufacturers
    copy_wb[sheet_name]['A70'].value = f'Top Suppliers {month_year_abbr} MTD'  # Ex: Top Suppliers  Apr-23  MTD
    copy_wb[sheet_name]['D70'].value = f'Top Suppliers {month_year_abbr} YTD'  # Ex: Top Suppliers  Apr-23  YTD

    copy_wb[sheet_name]['G70'].value = f'Top Manufacturers {month_year_abbr} MTD'  # Ex: Top Manufacturers  Apr-23  MTD
    copy_wb[sheet_name]['J70'].value = f'Top Manufacturers {month_year_abbr} YTD'  # Ex: Top Manufacturers  Apr-23  YTD

    copy_wb[sheet_name]['A85'].value = f'Top Product Categories {month_year_abbr} MTD'  # Ex: Top Product Categories  Apr-23  MTD
    copy_wb[sheet_name]['G85'].value = f'Top Product Categories {month_year_abbr} YTD'  # Ex: Top Product Categories  Apr-23  YTD


    for run in range(len(direct_non_pharma_spend_top_manufacturers_value_mtd)): # Looping 10 Times 
        ####### Row 72 to 81 ######## 
        cell_number = 72 + run 
        
        ## A,B,C :  direct_non_pharma_spend_top_suppliers_value_mtd | cell_value_spend | cell_value_percentage | Row 72 to 81 | A,B,C
        ABC_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_suppliers_value_mtd, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)

        ## D,E,F :  direct_non_pharma_spend_top_suppliers_value_ytd | cell_value_spend | cell_value_percentage | Row 72 to 81 | D,E,F
        DEF_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_suppliers_value_ytd, 'D', 'E', 'F', copy_wb, cell_number, sheet_name, run)
        
        ## G,H,I :  direct_non_pharma_spend_top_manufacturers_value_mtd | cell_value_spend | cell_value_percentage | Row 72 to 81 | G,H,I
        GHI_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_manufacturers_value_mtd, 'G', 'H', 'I', copy_wb, cell_number, sheet_name, run)
        
        ## J,K,L :  direct_non_pharma_spend_top_manufacturers_value_ytd | cell_value_spend | cell_value_percentage | Row 72 to 81 | J,K,L
        JKL_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_manufacturers_value_ytd, 'J', 'K', 'L', copy_wb, cell_number, sheet_name, run)
        
        
        ## Product Categories  | Row 87 - 92 | MTD | YTD | direct_non_pharma_spend_top_product_categories_value_mtd | direct_non_pharma_spend_top_product_categories_value_ytd
        start_value = 87
        cell_number = start_value + run 
        if cell_number<=start_value + 5 and run<=5:
            
            # A,B,C :  direct_non_pharma_spend_top_product_categories_value_mtd | cell_value_spend | cell_value_percentage | Row 87 - (87+5) | List Values from 0-5
            ABC_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_product_categories_value_mtd, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)
            
            # G,H,I :  direct_non_pharma_spend_top_product_categories_value_ytd | cell_value_spend | cell_value_percentage | Row 87 - (87+5) | List Values from 0-5
            GHI_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_product_categories_value_ytd, 'G', 'H', 'I', copy_wb, cell_number, sheet_name, run)
        
        new_run = run + 6
        if cell_number<=start_value + 4 and new_run < 10:
            
            # D,E,F :  direct_non_pharma_spend_top_product_categories_value_mtd | cell_value_spend | cell_value_percentage | Row 87 - (87+4) | List Values from 6-9
            DEF_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_product_categories_value_mtd, 'D', 'E', 'F', copy_wb, cell_number, sheet_name, new_run) # new_run
            
            # J,K,L :  direct_non_pharma_spend_top_product_categories_value_ytd | cell_value_spend | cell_value_percentage | Row 87 - (87+4) | List Values from 6-9
            JKL_Cell_values = spend_percentage_value_function(direct_non_pharma_spend_top_product_categories_value_ytd, 'J', 'K', 'L', copy_wb, cell_number, sheet_name, new_run) # new_run
        

    ## Providing All Others Info | B, C, E, F, H, I, K, L  | Row 82 | Top Manufacturers | Top Suppliers - 'All Others' Value
    all_others, all_others_percentage = getting_all_others_info(direct_non_pharma_spend_top_suppliers_value_mtd) # | Top Suppliers
    copy_wb[sheet_name]['B82'].value = all_others
    copy_wb[sheet_name]['C82'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_non_pharma_spend_top_suppliers_value_ytd) # | Top Suppliers
    copy_wb[sheet_name]['E82'].value = all_others
    copy_wb[sheet_name]['F82'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_non_pharma_spend_top_manufacturers_value_mtd) # | Top Manufacturers
    copy_wb[sheet_name]['H82'].value = all_others
    copy_wb[sheet_name]['I82'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_non_pharma_spend_top_manufacturers_value_ytd) # | Top Manufacturers
    copy_wb[sheet_name]['K82'].value = all_others
    copy_wb[sheet_name]['L82'].value = all_others_percentage

    ## Providing All Others Info | E, F, K, L | Row 91 | MTD | YTD | Top Product Categories 
    all_others, all_others_percentage = getting_all_others_info(direct_non_pharma_spend_top_product_categories_value_mtd) 
    copy_wb[sheet_name]['E91'].value = all_others
    copy_wb[sheet_name]['F91'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_non_pharma_spend_top_product_categories_value_ytd) 
    copy_wb[sheet_name]['K91'].value = all_others
    copy_wb[sheet_name]['L91'].value = all_others_percentage
    # =_=_==_=_== {7} Direct Pharma Spend | From [ROW-111]  =_=_==_=_==
    # {7.1} Total P.O. Spend - MTD [ROW-62]
    query_5_1_po_spend = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_5_1_po_spend)
    direct_pharma_spend_total_po_spend = curr.fetchall()

    # {7.2} MTD USD Spend [ROW-62] -------- No Need ---------- Taken from Excel Formula
    # {7.3} YTD USD Spend [ROW-62]  ------------- No Need ------------ Taken from Excel Formula


    # {7.4} Total P.O. Spend - YTD Trended [ROW -73]  | {7.4} Total P.O. Spend - YTD Trended [ROW -73] 

    ## 7.4.1 April YTD USD [ROW -74]
    query_5_4_1_ytd = f"""select geography ,sum (original_total_cost) as MTD_April_spend ,(MTD_April_spend / SUM(MTD_April_spend) OVER ()) * 100 AS percentage_MTD_April_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_5_4_1_ytd)
    direct_pharma_spend_ytd  = curr.fetchall()

    ## 7.4.2 JAN -23 [ROW -74] | three_months_prior  
    starting_date, ending_date = three_months_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_pharma_spend_three_months_prior_value = curr.fetchall()


    ## 7.4.3 FEB -23 [ROW-74] |  two_months_prior  
    starting_date, ending_date = two_months_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_pharma_spend_two_months_prior_value = curr.fetchall()


    ## 7.4.4 MAR -23 [ROW-74] |  one_month_prior  
    starting_date, ending_date = one_month_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_pharma_spend_one_month_prior_value = curr.fetchall()


    ## 7.4.5 APR - 23 [ROW -74] | current_date 
    starting_date, ending_date = current_date[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    direct_pharma_spend_current_date_value = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Direct (Pharma) Spend')

    # Writing in Cells - B99, B110 | Headers
    copy_wb[sheet_name]['B99'].value = f'{month_year_abbr} ( {curr_symbol} )'  # Ex: Apr-23 ( USD )
    copy_wb[sheet_name]['B110'].value = f"{month_year_abbr.split('-')[0]} YTD {curr_symbol}"  # Ex: Apr YTD USD 

    # Writing in Cells - D110, E110, F110, G110 | Headers 
    copy_wb[sheet_name]['D110'].value = three_months_prior[-1]  # Jan-23
    copy_wb[sheet_name]['E110'].value = two_months_prior[-1]  # Feb-23
    copy_wb[sheet_name]['F110'].value = one_month_prior[-1]  # Mar-23
    copy_wb[sheet_name]['G110'].value = current_date[-1]  # Apr-23

    index_to_add = market_df[market_df['country'] == selected_country].index[0] - 1

    ####### Row 100 to 103 [One Row] ######## 
    run = 0
    cell_number = 100 + index_to_add

    ## B:  direct_pharma_spend_total_po_spend | cell_value_spend | cell_value_percentage | Row 100 to 103 [One Row]
    B_Cell_values = spend_percentage_value_function(direct_pharma_spend_total_po_spend, False, 'B', False, copy_wb, cell_number, sheet_name, run)


    ####### Row 111 to 114 [One Row] ######## 
    cell_number = 111 + index_to_add
    ## B:  direct_pharma_spend_ytd | cell_value_spend | cell_value_percentage | Row 111 to 114 [One Row]
    B_Cell_values = spend_percentage_value_function(direct_pharma_spend_ytd, False, 'B', False, copy_wb, cell_number, sheet_name, run)

    ## D:  direct_pharma_spend_three_months_prior_value | cell_value_spend | Row 111 to 114 [One Row]
    D_Cell_values = spend_percentage_value_function(direct_pharma_spend_three_months_prior_value, False, 'D', False, copy_wb, cell_number, sheet_name, run)

    ## E:  direct_pharma_spend_two_months_prior_value | cell_value_spend | Row 111 to 114 [One Row]
    E_Cell_values = spend_percentage_value_function(direct_pharma_spend_two_months_prior_value, False, 'E', False, copy_wb, cell_number, sheet_name, run)

    ## F:  direct_pharma_spend_one_month_prior_value | cell_value_spend | Row 111 to 114 [One Row]
    F_Cell_values = spend_percentage_value_function(direct_pharma_spend_one_month_prior_value, False, 'F', False, copy_wb, cell_number, sheet_name, run)

    ## G:  direct_pharma_spend_current_date_value | cell_value_spend | Row 111 to 114 [One Row]
    G_Cell_values = spend_percentage_value_function(direct_pharma_spend_current_date_value, False, 'G', False, copy_wb, cell_number, sheet_name, run)
    # =_=_==_=_== {8} Top Manufacturers | Top Suppliers | Top Product Categories | Direct (Pharma) Spend=_=_==_=_==
    # {8.1} Top Manufacturers  Apr-23  MTD | Row 123
    query_top_manufacturers = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_manufacturers)
    direct_pharma_spend_top_manufacturers_value_mtd = curr.fetchall()

    # {8.2} Top Manufacturers  Apr-23  YTD | Row 123
    query_top_manufacturers_ytd = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_manufacturers_ytd)
    direct_pharma_spend_top_manufacturers_value_ytd = curr.fetchall()

    # {8.3} Top Suppliers  Apr-23  MTD | Row 123
    query_top_suppliers = f"""select distributor_normalized as suppliers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_suppliers)
    direct_pharma_spend_top_suppliers_value_mtd = curr.fetchall()

    # {8.4} Top Suppliers  Apr-23  YTD | Row 123
    query_top_suppliers_ytd = f"""select distributor_normalized as suppliers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_suppliers_ytd)
    direct_pharma_spend_top_suppliers_value_ytd = curr.fetchall()

    # {8.5} Top Product Categories  Apr-23  MTD | Row 139
    query_top_product_categories = f"""select unspsc_class_title as categories ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_product_categories)
    direct_pharma_spend_top_product_categories_value_mtd = curr.fetchall()

    # {8.6} Top Product Categories  Apr-23  YTD	| Row 139			
    query_top_product_categories_ytd = f"""select unspsc_class_title as categories ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_product_categories_ytd)
    direct_pharma_spend_top_product_categories_value_ytd = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Direct (Pharma) Spend | Top Suppliers | Top Manufacturers | Top Product Categories')

    # Writing in Cells - A70, D70, G70, J70, A85, G85 | Headers | Suppliers, Manufacturers
    copy_wb[sheet_name]['A123'].value = f'Top Suppliers {month_year_abbr} MTD'  # Ex: Top Suppliers  Apr-23  MTD
    copy_wb[sheet_name]['D123'].value = f'Top Suppliers {month_year_abbr} YTD'  # Ex: Top Suppliers  Apr-23  YTD

    copy_wb[sheet_name]['G123'].value = f'Top Manufacturers {month_year_abbr} MTD'  # Ex: Top Manufacturers  Apr-23  MTD
    copy_wb[sheet_name]['J123'].value = f'Top Manufacturers {month_year_abbr} YTD'  # Ex: Top Manufacturers  Apr-23  YTD

    copy_wb[sheet_name]['A139'].value = f'Top Product Categories {month_year_abbr} MTD'  # Ex: Top Product Categories  Apr-23  MTD
    copy_wb[sheet_name]['G139'].value = f'Top Product Categories {month_year_abbr} YTD'  # Ex: Top Product Categories  Apr-23  YTD


    for run in range(len(direct_pharma_spend_top_manufacturers_value_mtd)): # Looping 10 Times 
        ####### Row 125 to 134 ######## 
        cell_number = 125 + run 
        
        ## A,B,C :  direct_pharma_spend_top_suppliers_value_mtd | cell_value_spend | cell_value_percentage | Row 125 to 134 | A,B,C
        ABC_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_suppliers_value_mtd, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)

        ## D,E,F :  direct_pharma_spend_top_suppliers_value_ytd | cell_value_spend | cell_value_percentage | Row 125 to 134 | D,E,F
        DEF_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_suppliers_value_ytd, 'D', 'E', 'F', copy_wb, cell_number, sheet_name, run)
        
        ## G,H,I :  direct_pharma_spend_top_manufacturers_value_mtd | cell_value_spend | cell_value_percentage | Row 125 to 134 | G,H,I
        GHI_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_manufacturers_value_mtd, 'G', 'H', 'I', copy_wb, cell_number, sheet_name, run)
        
        ## J,K,L :  direct_pharma_spend_top_manufacturers_value_ytd | cell_value_spend | cell_value_percentage | Row 125 to 134 | J,K,L
        JKL_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_manufacturers_value_ytd, 'J', 'K', 'L', copy_wb, cell_number, sheet_name, run)
        
        
        ## Product Categories  | Row 141 to 146 | MTD | YTD | direct_pharma_spend_top_product_categories_value_mtd | direct_pharma_spend_top_product_categories_value_ytd
        start_value = 141
        cell_number = start_value + run 
        if cell_number<=start_value + 5 and run<=5:
            
            # A,B,C :  direct_pharma_spend_top_product_categories_value_mtd | cell_value_spend | cell_value_percentage | Row 87 - (87+5) | List Values from 0-5
            ABC_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_product_categories_value_mtd, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)
            
            # G,H,I :  direct_pharma_spend_top_product_categories_value_ytd | cell_value_spend | cell_value_percentage | Row 87 - (87+5) | List Values from 0-5
            GHI_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_product_categories_value_ytd, 'G', 'H', 'I', copy_wb, cell_number, sheet_name, run)
        
        new_run = run + 6
        if cell_number<=start_value + 4 and new_run < 10:
            
            # D,E,F :  direct_pharma_spend_top_product_categories_value_mtd | cell_value_spend | cell_value_percentage | Row 87 - (87+4) | List Values from 6-9
            DEF_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_product_categories_value_mtd, 'D', 'E', 'F', copy_wb, cell_number, sheet_name, new_run) # new_run
            
            # J,K,L :  direct_pharma_spend_top_product_categories_value_ytd | cell_value_spend | cell_value_percentage | Row 87 - (87+4) | List Values from 6-9
            JKL_Cell_values = spend_percentage_value_function(direct_pharma_spend_top_product_categories_value_ytd, 'J', 'K', 'L', copy_wb, cell_number, sheet_name, new_run) # new_run
        

    ## Providing All Others Info | B, C, E, F, H, I, K, L  | Row 135 | Top Manufacturers | Top Suppliers - 'All Others' Value
    all_others, all_others_percentage = getting_all_others_info(direct_pharma_spend_top_suppliers_value_mtd) # | Top Suppliers
    copy_wb[sheet_name]['B135'].value = all_others
    copy_wb[sheet_name]['C135'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_pharma_spend_top_suppliers_value_ytd) # | Top Suppliers
    copy_wb[sheet_name]['E135'].value = all_others
    copy_wb[sheet_name]['F135'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_pharma_spend_top_manufacturers_value_mtd) # | Top Manufacturers
    copy_wb[sheet_name]['H135'].value = all_others
    copy_wb[sheet_name]['I135'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_pharma_spend_top_manufacturers_value_ytd) # | Top Manufacturers
    copy_wb[sheet_name]['K135'].value = all_others
    copy_wb[sheet_name]['L135'].value = all_others_percentage

    ## Providing All Others Info | E, F, K, L | Row 145 | MTD | YTD | Top Product Categories 
    all_others, all_others_percentage = getting_all_others_info(direct_pharma_spend_top_product_categories_value_mtd) 
    copy_wb[sheet_name]['E145'].value = all_others
    copy_wb[sheet_name]['F145'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(direct_pharma_spend_top_product_categories_value_ytd) 
    copy_wb[sheet_name]['K145'].value = all_others
    copy_wb[sheet_name]['L145'].value = all_others_percentage
    # =_=_==_=_== {9} Indirect Spend | From [ROW-50]  =_=_==_=_==
    # {9.1} Total P.O. Spend - MTD [ROW-152]
    query_5_1_po_spend = f"""select geography ,sum (original_total_cost) as spend ,(spend / SUM(spend) OVER ()) * 100 AS percentage_spend from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_5_1_po_spend)
    indirect_spend_total_po_spend = curr.fetchall()

    # {9.2} MTD USD Spend [ROW-152] -------- No Need ---------- Taken from Excel Formula
    # {9.3} YTD USD Spend [ROW-152]  ------------- No Need ------------ Taken from Excel Formula


    # {9.4} Total P.O. Spend - YTD Trended [ROW - 152]  | {9.4} Total P.O. Spend - YTD Trended [ROW -73] 

    ## 9.4.1 April YTD USD [ROW -153]
    query_5_4_1_ytd = f"""select geography ,sum (original_total_cost) as MTD_April_spend ,(MTD_April_spend / SUM(MTD_April_spend) OVER ()) * 100 AS percentage_MTD_April_spend from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query_5_4_1_ytd)
    indirect_spend_ytd  = curr.fetchall()

    ## 9.4.2 JAN -23 [ROW -153] | three_months_prior  
    starting_date, ending_date = three_months_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    indirect_spend_three_months_prior_value = curr.fetchall()


    ## 9.4.3 FEB -23 [ROW-153] |  two_months_prior  
    starting_date, ending_date = two_months_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    indirect_spend_two_months_prior_value = curr.fetchall()


    ## 9.4.4 MAR -23 [ROW-153] |  one_month_prior  
    starting_date, ending_date = one_month_prior[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    indirect_spend_one_month_prior_value = curr.fetchall()


    ## 9.4.5 APR - 23 [ROW -153] | current_date 
    starting_date, ending_date = current_date[0:2]
    query = f"""select geography ,sum (original_total_cost) as spend from {table_name} where date_of_purchase > '{starting_date}' and date_of_purchase < '{ending_date}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1"""
    curr.execute(query)
    indirect_spend_current_date_value = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Indirect Spend')

    # Writing in Cells - B153, G153 | Headers
    copy_wb[sheet_name]['B153'].value = f'{month_year_abbr} ( {curr_symbol} )'  # Ex: Apr-23 ( USD )
    copy_wb[sheet_name]['G153'].value = f"{month_year_abbr.split('-')[0]} YTD {curr_symbol}"  # Ex: Apr YTD USD 

    # Writing in Cells - I153, J153, K153, L153 | Headers 
    copy_wb[sheet_name]['I153'].value = three_months_prior[-1]  # Jan-23
    copy_wb[sheet_name]['J153'].value = two_months_prior[-1]  # Feb-23
    copy_wb[sheet_name]['K153'].value = one_month_prior[-1]  # Mar-23
    copy_wb[sheet_name]['L153'].value = current_date[-1]  # Apr-23


    index_to_add = market_df[market_df['country'] == selected_country].index[0] - 1

    ####### Row 154 to 157 [One Row] ######## 
    run = 0
    cell_number = 154 + index_to_add

    ## B:  indirect_spend_total_po_spend | cell_value_spend | cell_value_percentage | Row 154 to 157 [One Row]
    B_Cell_values = spend_percentage_value_function(indirect_spend_total_po_spend, False, 'B', False, copy_wb, cell_number, sheet_name, run)

    ## G:  indirect_spend_ytd | cell_value_spend | cell_value_percentage | Row 154 to 157 [One Row]
    G_Cell_values = spend_percentage_value_function(indirect_spend_ytd, False, 'G', False, copy_wb, cell_number, sheet_name, run)

    ## I:  indirect_spend_three_months_prior_value | cell_value_spend | Row 154 to 157 [One Row]
    I_Cell_values = spend_percentage_value_function(indirect_spend_three_months_prior_value, False, 'I', False, copy_wb, cell_number, sheet_name, run)

    ## J:  indirect_spend_two_months_prior_value | cell_value_spend | Row 154 to 157 [One Row]
    J_Cell_values = spend_percentage_value_function(indirect_spend_two_months_prior_value, False, 'J', False, copy_wb, cell_number, sheet_name, run)

    ## K:  indirect_spend_one_month_prior_value | cell_value_spend | Row 154 to 157 [One Row]
    K_Cell_values = spend_percentage_value_function(indirect_spend_one_month_prior_value, False, 'K', False, copy_wb, cell_number, sheet_name, run)

    ## L:  indirect_spend_current_date_value | cell_value_spend | Row 154 to 157 [One Row]
    L_Cell_values = spend_percentage_value_function(indirect_spend_current_date_value, False, 'L', False, copy_wb, cell_number, sheet_name, run)
    # =_=_==_=_== {10} Top Manufacturers | Top Suppliers | Top Product Categories | Indirect Spend=_=_==_=_==
    # {10.1} Top Manufacturers  Apr-23  MTD | Row 163
    query_top_manufacturers = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_manufacturers)
    indirect_spend_top_manufacturers_value_mtd = curr.fetchall()

    # {10.2} Top Manufacturers  Apr-23  YTD | Row 163
    query_top_manufacturers_ytd = f"""select mnf_dashboard_half as manufacturers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_manufacturers_ytd)
    indirect_spend_top_manufacturers_value_ytd = curr.fetchall()

    # {10.3} Top Suppliers  Apr-23  MTD | Row 163
    query_top_suppliers = f"""select distributor_normalized as suppliers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_suppliers)
    indirect_spend_top_suppliers_value_mtd = curr.fetchall()

    # {10.4} Top Suppliers  Apr-23  YTD | Row 163
    query_top_suppliers_ytd = f"""select distributor_normalized as suppliers ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_suppliers_ytd)
    indirect_spend_top_suppliers_value_ytd = curr.fetchall()

    # {10.5} Top Product Categories  Apr-23  MTD | Row 178
    query_top_product_categories = f"""select unspsc_class_title as categories ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_mtd}' and date_of_purchase < '{end_date_mtd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_product_categories)
    indirect_spend_top_product_categories_value_mtd = curr.fetchall()

    # {10.6} Top Product Categories  Apr-23  YTD	| Row 178			
    query_top_product_categories_ytd = f"""select unspsc_class_title as categories ,sum (original_total_cost) as USD_global ,(USD_global/ SUM(USD_global) OVER ()) * 100 AS percentage_USD_global from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}' GROUP BY 1 ORDER BY 2 desc limit 10"""
    curr.execute(query_top_product_categories_ytd)
    indirect_spend_top_product_categories_value_ytd = curr.fetchall()
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Indirect Spend | Top Suppliers | Top Manufacturers | Top Product Categories')

    # Writing in Cells - A163, D163, G163, J163, A178, G178 | Headers | Suppliers, Manufacturers, Product Categories
    copy_wb[sheet_name]['A163'].value = f'Top Suppliers {month_year_abbr} MTD'  # Ex: Top Suppliers  Apr-23  MTD
    copy_wb[sheet_name]['D163'].value = f'Top Suppliers {month_year_abbr} YTD'  # Ex: Top Suppliers  Apr-23  YTD

    copy_wb[sheet_name]['G163'].value = f'Top Manufacturers {month_year_abbr} MTD'  # Ex: Top Manufacturers  Apr-23  MTD
    copy_wb[sheet_name]['J163'].value = f'Top Manufacturers {month_year_abbr} YTD'  # Ex: Top Manufacturers  Apr-23  YTD

    copy_wb[sheet_name]['A178'].value = f'Top Product Categories {month_year_abbr} MTD'  # Ex: Top Product Categories  Apr-23  MTD
    copy_wb[sheet_name]['G178'].value = f'Top Product Categories {month_year_abbr} YTD'  # Ex: Top Product Categories  Apr-23  YTD


    for run in range(len(indirect_spend_top_manufacturers_value_mtd)): # Looping 10 Times 
        ####### Row 165 to 174 ######## 
        cell_number = 165 + run 
        
        ## A,B,C :  indirect_spend_top_suppliers_value_mtd | cell_value_spend | cell_value_percentage | Row 165 to 174 | A,B,C
        ABC_Cell_values = spend_percentage_value_function(indirect_spend_top_suppliers_value_mtd, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)

        ## D,E,F :  indirect_spend_top_suppliers_value_ytd | cell_value_spend | cell_value_percentage | Row 165 to 174 | D,E,F
        DEF_Cell_values = spend_percentage_value_function(indirect_spend_top_suppliers_value_ytd, 'D', 'E', 'F', copy_wb, cell_number, sheet_name, run)
        
        ## G,H,I :  indirect_spend_top_manufacturers_value_mtd | cell_value_spend | cell_value_percentage | Row 165 to 174 | G,H,I
        GHI_Cell_values = spend_percentage_value_function(indirect_spend_top_manufacturers_value_mtd, 'G', 'H', 'I', copy_wb, cell_number, sheet_name, run)
        
        ## J,K,L :  indirect_spend_top_manufacturers_value_ytd | cell_value_spend | cell_value_percentage | Row 165 to 174 | J,K,L
        JKL_Cell_values = spend_percentage_value_function(indirect_spend_top_manufacturers_value_ytd, 'J', 'K', 'L', copy_wb, cell_number, sheet_name, run)
        
        
        ## Product Categories  | Row 180 to 185 | MTD | YTD | indirect_spend_top_product_categories_value_mtd | indirect_spend_top_product_categories_value_ytd
        start_value = 180
        cell_number = start_value + run 
        if cell_number<=start_value + 5 and run<=5:
            
            # A,B,C :  indirect_spend_top_product_categories_value_mtd | cell_value_spend | cell_value_percentage | Row 180 - (180+5) | List Values from 0-5
            ABC_Cell_values = spend_percentage_value_function(indirect_spend_top_product_categories_value_mtd, 'A', 'B', 'C', copy_wb, cell_number, sheet_name, run)
            
            # G,H,I :  indirect_spend_top_product_categories_value_ytd | cell_value_spend | cell_value_percentage | Row 180 - (180+5) | List Values from 0-5
            GHI_Cell_values = spend_percentage_value_function(indirect_spend_top_product_categories_value_ytd, 'G', 'H', 'I', copy_wb, cell_number, sheet_name, run)
        
        new_run = run + 6
        if cell_number<=start_value + 4 and new_run < 10:
            
            # D,E,F :  indirect_spend_top_product_categories_value_mtd | cell_value_spend | cell_value_percentage | Row 180 - (180+4) | List Values from 6-9
            DEF_Cell_values = spend_percentage_value_function(indirect_spend_top_product_categories_value_mtd, 'D', 'E', 'F', copy_wb, cell_number, sheet_name, new_run) # new_run
            
            # J,K,L :  indirect_spend_top_product_categories_value_ytd | cell_value_spend | cell_value_percentage | Row 180 - (180+4) | List Values from 6-9
            JKL_Cell_values = spend_percentage_value_function(indirect_spend_top_product_categories_value_ytd, 'J', 'K', 'L', copy_wb, cell_number, sheet_name, new_run) # new_run
        

    ## Providing All Others Info | B, C, E, F, H, I, K, L  | Row 175 | Top Manufacturers | Top Suppliers - 'All Others' Value
    all_others, all_others_percentage = getting_all_others_info(indirect_spend_top_suppliers_value_mtd) # | Top Suppliers
    copy_wb[sheet_name]['B175'].value = all_others
    copy_wb[sheet_name]['C175'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(indirect_spend_top_suppliers_value_ytd) # | Top Suppliers
    copy_wb[sheet_name]['E175'].value = all_others
    copy_wb[sheet_name]['F175'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(indirect_spend_top_manufacturers_value_mtd) # | Top Manufacturers
    copy_wb[sheet_name]['H175'].value = all_others
    copy_wb[sheet_name]['I175'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(indirect_spend_top_manufacturers_value_ytd) # | Top Manufacturers
    copy_wb[sheet_name]['K175'].value = all_others
    copy_wb[sheet_name]['L175'].value = all_others_percentage

    ## Providing All Others Info | E, F, K, L | Row 184 | MTD | YTD | Top Product Categories 
    all_others, all_others_percentage = getting_all_others_info(indirect_spend_top_product_categories_value_mtd) 
    copy_wb[sheet_name]['E184'].value = all_others
    copy_wb[sheet_name]['F184'].value = all_others_percentage

    all_others, all_others_percentage = getting_all_others_info(indirect_spend_top_product_categories_value_ytd) 
    copy_wb[sheet_name]['K184'].value = all_others
    copy_wb[sheet_name]['L184'].value = all_others_percentage
    # =_=_==_=_== {11} Total P.O. Spend - YoY Trend | H,J,L | Row: 21 to 25 =_=_==_=_==
    def get_date_ranges(year):
        year = int(year)
        date_ranges = []

        for i in range(year-3, year):
            date_range_start = datetime(i-1, 12, 31)  # End of the current year
            date_range_end = datetime(i+1, 1, 1)  # Start of the next year
            date_ranges.append((date_range_start.strftime('%Y-%m-%d'), date_range_end.strftime('%Y-%m-%d')))

        return date_ranges

    yoy_date_range = get_date_ranges(year)
    index_to_add = market_df[market_df['country'] == selected_country].index[0] - 1
    column_names_yoy_table = [yoy_date_range[1][0].split('-')[0], yoy_date_range[-1][0].split('-')[0], yoy_date_range[1][1].split('-')[0]]
    yoy_countries = ['Brazil', 'Chile', 'Colombia', 'Peru', 'Portugal']
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Total P.O. Spend - YoY Trend')

    total_po_spend_yoy_values = []
    for dt in yoy_date_range:
        start_yoy_date, end_yoy_date = dt
        
        if start_yoy_date == '2019-12-31' and end_yoy_date == '2021-01-01': # Query for 2020 Data
            query = f"""select sum (original_total_cost) as MTD_usd from {table_name} where geography = '{selected_country}' and date_of_purchase > '{start_yoy_date}' and date_of_purchase < '{end_yoy_date}' and lower(spend_type_1) = 'overall'"""
        else:
            query = f"""select sum (original_total_cost) as MTD_usd_local from {table_name} where date_of_purchase > '{start_yoy_date}' and date_of_purchase < '{end_yoy_date}' and geography = '{selected_country}'"""
        
        curr.execute(query)
        value = curr.fetchall()
        value = list(value[0].values())[0]
        total_po_spend_yoy_values.append(value)


    # H, J, L | Headers: Row 20 | Value: Row 21 | Ex: 2020 ( CLP )

    row1 = 20
    row2 = 21
    columns = ['H', 'J', 'L']
    for index, name in enumerate(column_names_yoy_table):
        cell_value = f'{name} ( {curr_symbol} )'
        cell_name = f'{columns[index]}{row1}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## Writing Values | total_po_spend_yoy_values
        cell_value = total_po_spend_yoy_values[index]
        cell_name = f'{columns[index]}{row2}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
    # =_=_==_=_== {12} Non-Pharma Spend YoY Trend | Headers: Row 30, 48 | Ex: 2020 ( CLP ) | Value: Row 31, 49-52 =_=_==_=_==
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Non-Pharma Spend YoY Trend')

    total_non_pharma_spend_yoy_values = []
    for dt in yoy_date_range:
        start_yoy_date, end_yoy_date = dt
        query = f"""select sum (original_total_cost) as MTD_usd_local from {table_name} where date_of_purchase > '{start_yoy_date}' and date_of_purchase < '{end_yoy_date}' and geography = '{selected_country}' and spend_type_1 = 'Direct (Non-Pharma)'"""
        curr.execute(query)
        value = curr.fetchall()
        value = list(value[0].values())[0]
        total_non_pharma_spend_yoy_values.append(value)
        
    # B, D, F, H, J, L | Headers: Row 30, 48 | Value: Row 31, 49-52 | Ex: 2020 ( CLP )

    columns = ['B', 'D', 'F']
    row1 = 30
    row2 = 31

    columns_2 = ['H', 'J', 'L']
    row3 = 48
    row4 = 49 + index_to_add

    for index, name in enumerate(column_names_yoy_table):
        
        ## B, D, F | Headers: Row 30| Ex: 2020 ( CLP ) 
        cell_value = f'{name} ( {curr_symbol} )'
        
        cell_name = f'{columns[index]}{row1}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## H, J, L | Headers: Row 48 | Ex: 2020 ( CLP ) 
        cell_name = f'{columns_2[index]}{row3}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## Writing Values | total_non_pharma_spend_yoy_values | Row 31, 49-52
        cell_value = total_non_pharma_spend_yoy_values[index]
        
        cell_name = f'{columns[index]}{row2}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## Writing Values | total_non_pharma_spend_yoy_values | 49-52
        cell_name = f'{columns_2[index]}{row4}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
    # =_=_==_=_== {13} Pharma Spend YoY Trend | Headers: Row 30, 99 | Ex: 2020 ( CLP ) | Value: Row 31, 100-103 =_=_==_=_==
    print(f'{selected_country} - {sheet_name} - {curr_symbol} | Table: Pharma Spend YoY Trend')

    total_pharma_spend_yoy_values = []
    for dt in yoy_date_range:
        start_yoy_date, end_yoy_date = dt
        query = f"""select sum (original_total_cost) as MTD_usd_local from {table_name} where date_of_purchase > '{start_yoy_date}' and date_of_purchase < '{end_yoy_date}' and geography = '{selected_country}' and spend_type_1 = 'Direct (Pharma)'"""
        curr.execute(query)
        value = curr.fetchall()
        value = list(value[0].values())[0]
        total_pharma_spend_yoy_values.append(value)
        
    # H, J, L | Headers: Row 30, 99 | Value: Row 31, 100-103 | Ex: 2020 ( CLP )

    columns = ['H', 'J', 'L']
    row1 = 30
    row2 = 31

    row3 = 99
    row4 = 100 + index_to_add

    for index, name in enumerate(column_names_yoy_table):
        
        ## H, J, L | Headers: Row 30| Ex: 2020 ( CLP ) 
        cell_value = f'{name} ( {curr_symbol} )'
        
        cell_name = f'{columns[index]}{row1}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## H, J, L | Headers: Row 99 | Ex: 2020 ( CLP ) 
        cell_name = f'{columns_2[index]}{row3}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## Writing Values | total_pharma_spend_yoy_values | Row 31
        cell_value = total_pharma_spend_yoy_values[index]
        
        cell_name = f'{columns[index]}{row2}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
        ## Writing Values | total_pharma_spend_yoy_values | Row 100-103
        cell_name = f'{columns_2[index]}{row4}'
        copy_wb[sheet_name][cell_name].value = cell_value
        # print(f'{cell_name} | {cell_value}')
        
    # Single Values |  A,E,I | Row 44 | *Total unique mfg  3131	| *Total unique mfg  693	| *Total unique mfg  518
    ### Ex: *Total unique mfg  3131 | A | Non-Pharma

    query = f"""select count(distinct(mnf_dashboard_half)) as total_unique_mfg_for_NonPharma from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Non-Pharma)' and geography = '{selected_country}'"""
    curr.execute(query)
    value = curr.fetchall()
    # print(value)
    copy_wb[sheet_name]['A44'].value = f"*Total unique mfg {list(value[0].values())[0]}"

    ### Ex: *Total unique mfg  693 | E | Pharma

    query = f"""select count(distinct(mnf_dashboard_half)) as total_unique_mfg_for_Pharma from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Direct (Pharma)' and geography = '{selected_country}'"""
    curr.execute(query)
    value = curr.fetchall()
    # print(value)
    copy_wb[sheet_name]['E44'].value = f"*Total unique mfg {list(value[0].values())[0]}"

    ### Ex: *Total unique mfg  518 | I | Indirect

    query = f"""select count(distinct(mnf_dashboard_half)) as total_unique_mfg_for_NonPharma from {table_name} where date_of_purchase > '{start_date_ytd}' and date_of_purchase < '{end_date_ytd}' and spend_type_1 = 'Indirect' and geography = '{selected_country}'"""
    curr.execute(query)
    value = curr.fetchall()
    # print(value)
    copy_wb[sheet_name]['I44'].value = f"*Total unique mfg {list(value[0].values())[0]}"
    # Saving Everything | Closing
    ## Adding Image
    img = Image('Template\image.png')
    sheet = copy_wb[sheet_name] 
    sheet.add_image(img, 'B8') 

    ## Removing Chart Title
    for chart in copy_wb[sheet_name]._charts:
        chart.title = None

    copy_wb.save(copy_file_path)
    print(f'Completed: {file_name}')